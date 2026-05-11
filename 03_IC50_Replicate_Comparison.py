#!/usr/bin/env python3
"""
IC50 Biological Replicate Comparison
Assesses day-to-day reproducibility across biological replicates for MG1655 and ACE-1.
Author: Felix Knote
"""

import os
import re
import warnings

# ─────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────
BASE_DIR       = r"L:\43-RVZ\AIMicroscopy\Mitarbeiter\3_Users\Felix\2026_04_24_AI4AB_IC50_Determination\Biological replicates"
PLATE_MAP_PATH = os.path.join(BASE_DIR, "AB Plate Map.csv")
STRAIN_COLORS  = {"MG1655": "#2166ac", "ACE-1": "#d6604d"}
REP_PALETTE    = ["#1b9e77", "#d95f02", "#7570b3", "#e7298a", "#66a61e"]
_TECH_COLOR    = "#5aafe0"   # blue – technical (within-day) variance
_BIO_COLOR     = "#e07b5a"   # orange – biological (day-to-day) variance
OUTPUT_DIR     = os.path.join(BASE_DIR, "Reproducibility")
# ─────────────────────────────────────────────────────────────

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from openpyxl.styles import PatternFill
from scipy.optimize import curve_fit
from scipy import stats

# Plate geometry
N_DILUTIONS       = 10
DILUTION_COLS     = list(range(2, 12))
BLANK_WELLS       = ["A1", "A12", "H1", "H12"]
GROWTH_CTRL_WELLS = [f"A{c}" for c in DILUTION_COLS]
AB1_ROWS          = list("BCD")
AB2_ROWS          = list("EFG")

_RAW_IC50_RENAME = {
    "Rep1_IC50":   "Rep 1 IC50 (µg/mL)",
    "Rep2_IC50":   "Rep 2 IC50 (µg/mL)",
    "Rep3_IC50":   "Rep 3 IC50 (µg/mL)",
    "Mean_IC50":   "Mean IC50 (µg/mL)",
    "SD_IC50":     "SD IC50 (µg/mL)",
    "Fold_change": "Max/Min fold change",
}


# ─────────────────────────────────────────────────────────────
# Data loading & parsing
# ─────────────────────────────────────────────────────────────
_WELL_LINE = re.compile(r'^([A-Ha-h])(\d{1,2})\s*;\s*([0-9.,\-]+)')


def _to_float(val) -> float:
    if pd.isna(val):
        return np.nan
    s = str(val).strip().replace(" ", "")
    if s.count(",") == 1 and s.count(".") >= 1 and s.index(",") > s.rindex("."):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return np.nan


def load_plate(filepath: str) -> pd.DataFrame:
    records = []
    with open(filepath, "r", encoding="utf-8", errors="replace") as fh:
        for line in fh:
            m = _WELL_LINE.match(line.strip())
            if m:
                records.append({
                    "Well": m.group(1).upper() + str(int(m.group(2))),
                    "OD":   _to_float(m.group(3)),
                })
    if not records:
        raise ValueError(f"No well data found in: {filepath}")
    df = pd.DataFrame(records)
    if len(df) != 96:
        warnings.warn(f"{os.path.basename(filepath)}: expected 96 wells, got {len(df)}")
    return df


def load_plate_map(filepath: str) -> pd.DataFrame:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        raw = None
        for sep in [",", ";", "\t"]:
            try:
                candidate = pd.read_csv(filepath, sep=sep)
                if candidate.shape[1] >= 3:
                    raw = candidate
                    break
            except Exception:
                continue
        if raw is None:
            raw = pd.read_csv(filepath)
    else:
        raw = pd.read_excel(filepath)
    raw.columns = [c.strip() for c in raw.columns]
    name_col, conc_col = raw.columns[1], raw.columns[2]
    records = []
    for i in range(0, len(raw), 2):
        ab1 = raw.iloc[i]
        ab2 = raw.iloc[i + 1] if (i + 1) < len(raw) else None
        records.append({
            "plate_id":              i // 2 + 1,
            "antibiotic_1_name":     str(ab1[name_col]).strip(),
            "antibiotic_1_max_conc": float(ab1[conc_col]),
            "antibiotic_2_name":     str(ab2[name_col]).strip() if ab2 is not None else None,
            "antibiotic_2_max_conc": float(ab2[conc_col])      if ab2 is not None else None,
        })
    return pd.DataFrame(records)


def parse_plate_layout(plate_df: pd.DataFrame):
    def wells_od(wells):
        return plate_df.loc[plate_df["Well"].isin(wells), "OD"].values.astype(float)
    blank_od  = float(np.nanmean(wells_od(BLANK_WELLS)))
    growth_od = wells_od(GROWTH_CTRL_WELLS)
    ab1_od    = {row: wells_od([f"{row}{c}" for c in DILUTION_COLS]) for row in AB1_ROWS}
    ab2_od    = {row: wells_od([f"{row}{c}" for c in DILUTION_COLS]) for row in AB2_ROWS}
    return blank_od, growth_od, ab1_od, ab2_od


def normalize_data(raw_od: np.ndarray, blank_od: float, growth_od: np.ndarray) -> np.ndarray:
    corrected        = np.clip(raw_od    - blank_od, 0, None)
    growth_corrected = np.clip(growth_od - blank_od, 0, None)
    norm_factor = float(np.nanmean(growth_corrected))
    if norm_factor == 0:
        warnings.warn("Growth control mean is zero – normalization skipped.")
        return corrected
    return corrected / norm_factor * 100.0


# ─────────────────────────────────────────────────────────────
# Dose-response fitting (4PL)
# ─────────────────────────────────────────────────────────────
def four_param_logistic(x, bottom, top, log_ic50, hill):
    return bottom + (top - bottom) / (1.0 + 10.0 ** ((log_ic50 - np.log10(x)) * hill))


def fit_dose_response(concentrations: np.ndarray, responses: np.ndarray):
    mask = ~np.isnan(responses) & (concentrations > 0)
    x, y = concentrations[mask], responses[mask]
    if len(x) < 4:
        raise RuntimeError("Fewer than 4 valid data points.")
    lx = np.log10(x)
    lb = [  0.0,   0.0, lx.min() - 1.0,  0.1]
    ub = [200.0, 200.0, lx.max() + 1.0, 10.0]
    p0 = [
        np.clip(y.min(),       lb[0] + 1e-9, ub[0] - 1e-9),
        np.clip(y.max(),       lb[1] + 1e-9, ub[1] - 1e-9),
        np.clip(np.median(lx), lb[2] + 1e-9, ub[2] - 1e-9),
        np.clip( 1.0,          lb[3] + 1e-9, ub[3] - 1e-9),
    ]
    popt, pcov = curve_fit(  # type: ignore[misc]
        four_param_logistic, x, y,
        p0=p0, bounds=(lb, ub), maxfev=50_000, method="trf",
    )
    y_pred = four_param_logistic(x, *popt)
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - y.mean()) ** 2)
    r2     = float(1 - ss_res / ss_tot) if ss_tot > 0 else np.nan
    return 10.0 ** popt[2], popt, r2, pcov


def calculate_ic50(rep_responses: dict, concentrations: np.ndarray):
    results = []
    for rep, resp in rep_responses.items():
        try:
            ic50, popt, r2, pcov = fit_dose_response(concentrations, resp)
            _, _, log_ic50, _ = popt
            log_ic50_se = float(np.sqrt(pcov[2, 2])) if np.isfinite(pcov[2, 2]) else np.nan
            ic50_ci_lo  = 10.0 ** (log_ic50 - 1.96 * log_ic50_se) if not np.isnan(log_ic50_se) else np.nan
            ic50_ci_hi  = 10.0 ** (log_ic50 + 1.96 * log_ic50_se) if not np.isnan(log_ic50_se) else np.nan
            results.append({
                "rep": rep, "ic50": ic50, "popt": popt, "r2": r2, "responses": resp,
                "ic50_ci_lo": ic50_ci_lo, "ic50_ci_hi": ic50_ci_hi,
            })
        except Exception as exc:
            warnings.warn(f"Fit failed for {rep}: {exc}")
            results.append({
                "rep": rep, "ic50": np.nan, "popt": None, "r2": np.nan, "responses": resp,
                "ic50_ci_lo": np.nan, "ic50_ci_hi": np.nan,
            })
    ic50_vals = [r["ic50"] for r in results if not np.isnan(r["ic50"])]
    mean_ic50 = float(np.mean(ic50_vals))           if ic50_vals          else np.nan
    sd_ic50   = float(np.std(ic50_vals, ddof=1))    if len(ic50_vals) > 1 else np.nan
    return results, mean_ic50, sd_ic50


# ─────────────────────────────────────────────────────────────
# Filename discovery
# ─────────────────────────────────────────────────────────────
# Matches e.g. "260422_FK OD600 after 18h_2026_04_22_FK_IC50_MG1655_P1.csv"
#          and "260424_FK OD600 after 18h_2026_04_24_FK_ACE-1_P3.csv"
_FNAME_RE = re.compile(
    r'(\d{4}_\d{2}_\d{2})_FK(?:_IC50)?_([A-Za-z0-9][A-Za-z0-9-]*)_P(\d+)\.csv$',
    re.IGNORECASE,
)


def parse_filename(fname: str):
    """Returns (date_str, strain, plate_id) or None."""
    m = _FNAME_RE.search(fname)
    if not m:
        return None
    return m.group(1).replace("_", "-"), m.group(2), int(m.group(3))


def discover_files(base_dir: str) -> dict:
    """Returns {strain: {date: [(plate_id, filepath)]}}"""
    result = {}
    for fname in sorted(os.listdir(base_dir)):
        parsed = parse_filename(fname)
        if parsed is None:
            continue
        date, strain, plate_id = parsed
        result.setdefault(strain, {}).setdefault(date, []).append(
            (plate_id, os.path.join(base_dir, fname))
        )
    return result


# ─────────────────────────────────────────────────────────────
# Process one (strain, date) experiment
# ─────────────────────────────────────────────────────────────
def process_experiment(plate_files: list, plate_map: pd.DataFrame) -> list:
    results = []
    for plate_id, fpath in sorted(plate_files):
        map_row = plate_map[plate_map["plate_id"] == plate_id]
        if map_row.empty:
            continue
        map_row = map_row.iloc[0]
        try:
            plate_df = load_plate(fpath)
        except Exception as exc:
            warnings.warn(f"Could not load {os.path.basename(fpath)}: {exc}")
            continue
        blank_od, growth_od, ab1_od, ab2_od = parse_plate_layout(plate_df)
        antibiotics = []
        if map_row["antibiotic_1_name"]:
            antibiotics.append((map_row["antibiotic_1_name"],
                                 float(map_row["antibiotic_1_max_conc"]), ab1_od))
        if map_row["antibiotic_2_name"]:
            antibiotics.append((map_row["antibiotic_2_name"],
                                 float(map_row["antibiotic_2_max_conc"]), ab2_od))
        for ab_name, ab_max_conc, ab_od in antibiotics:
            conc    = np.array([ab_max_conc / 2**i for i in range(N_DILUTIONS)])
            ab_norm = {f"Rep {i+1}": normalize_data(od, blank_od, growth_od)
                       for i, od in enumerate(ab_od.values())}
            rep_results, mean_ic50, sd_ic50 = calculate_ic50(ab_norm, conc)
            ci_los     = [r["ic50_ci_lo"] for r in rep_results if not np.isnan(r["ic50_ci_lo"])]
            ci_his     = [r["ic50_ci_hi"] for r in rep_results if not np.isnan(r["ic50_ci_hi"])]
            results.append({
                "ab_name":        ab_name,
                "plate_id":       plate_id,
                "mean_ic50":      mean_ic50,
                "sd_ic50":        sd_ic50,
                "mean_ci_lo":     float(np.mean(ci_los)) if ci_los else np.nan,
                "mean_ci_hi":     float(np.mean(ci_his)) if ci_his else np.nan,
                "rep_ic50s":      [r["ic50"] for r in rep_results],
                "concentrations": conc,
                "rep_results":    rep_results,
            })
    return results


# ─────────────────────────────────────────────────────────────
# Reproducibility table
# ─────────────────────────────────────────────────────────────
def build_reproducibility_dfs(all_results: dict) -> dict:
    """
    all_results: {strain: {date: [result_dicts]}}
    Returns {strain: DataFrame} with per-date IC50s, mean, SD, CV%, log10_SD.
    Each row is one antibiotic; each biological replicate (date) gets its own IC50 and SD column.
    """
    strain_dfs = {}
    for strain, date_results in all_results.items():
        dates   = sorted(date_results.keys())
        ab_data = {}
        for date in dates:
            for r in date_results[date]:
                ab_data.setdefault(r["ab_name"], {})[date] = {
                    "mean_ic50": r["mean_ic50"],
                    "sd_ic50":   r["sd_ic50"],
                    "mean_ci_lo": r.get("mean_ci_lo", np.nan),
                    "mean_ci_hi": r.get("mean_ci_hi", np.nan),
                }
        rows = []
        for ab_name in sorted(ab_data):
            row  = {"Antibiotic": ab_name}
            vals = []
            for date in dates:
                entry = ab_data[ab_name].get(date, {})
                ic50  = entry.get("mean_ic50", np.nan)
                sd    = entry.get("sd_ic50",   np.nan)
                row[f"{date}_IC50"]   = ic50
                row[f"{date}_SD"]     = sd
                row[f"{date}_CI_lo"]  = entry.get("mean_ci_lo", np.nan)
                row[f"{date}_CI_hi"]  = entry.get("mean_ci_hi", np.nan)
                if not np.isnan(ic50):
                    vals.append(ic50)
            row["Mean_IC50"] = float(np.mean(vals))           if vals          else np.nan
            row["SD_IC50"]   = float(np.std(vals, ddof=1))   if len(vals) > 1 else np.nan
            mean_v = row["Mean_IC50"]
            sd_v   = row["SD_IC50"]
            row["CV_pct"]   = sd_v / mean_v * 100 if (mean_v and not np.isnan(sd_v)) else np.nan
            log_vals        = [np.log10(v) for v in vals if v > 0]
            row["log10_SD"] = float(np.std(log_vals, ddof=1)) if len(log_vals) > 1 else np.nan
            row["n_reps"]   = len(vals)
            rows.append(row)
        strain_dfs[strain] = pd.DataFrame(rows)
    return strain_dfs


def make_rep_colors(dates: list) -> dict:
    return {d: REP_PALETTE[i % len(REP_PALETTE)] for i, d in enumerate(sorted(dates))}


# ─────────────────────────────────────────────────────────────
# Plots
# ─────────────────────────────────────────────────────────────


def plot_cv_heatmap(strain_dfs: dict, save_dir: str) -> str:
    """
    Heatmap of CV% – rows = antibiotics, columns = strains.
    Green = reproducible (<20%), red = variable (>50%).
    """
    strains = list(strain_dfs.keys())
    all_abs = sorted(set().union(*[set(df["Antibiotic"]) for df in strain_dfs.values()]))

    mat = np.full((len(all_abs), len(strains)), np.nan)
    for j, strain in enumerate(strains):
        cv_map = strain_dfs[strain].set_index("Antibiotic")["CV_pct"].to_dict()
        for i, ab in enumerate(all_abs):
            mat[i, j] = cv_map.get(ab, np.nan)

    fig, ax = plt.subplots(figsize=(max(4, len(strains) * 2.0),
                                    max(6, len(all_abs) * 0.5)))
    im = ax.imshow(mat, aspect="auto", cmap="RdYlGn_r", vmin=0, vmax=50)
    plt.colorbar(im, ax=ax, label="CV% across biological replicates")
    ax.set_xticks(range(len(strains)))
    ax.set_xticklabels(strains, fontsize=11)
    ax.set_yticks(range(len(all_abs)))
    ax.set_yticklabels(all_abs, fontsize=9)
    for i in range(len(all_abs)):
        for j in range(len(strains)):
            val = mat[i, j]
            if not np.isnan(val):
                ax.text(j, i, f"{val:.0f}%", ha="center", va="center",
                        fontsize=8, color="black" if val < 35 else "white")
    ax.set_title("CV% Heatmap: Biological Replicate Reproducibility",
                 fontsize=13, fontweight="bold", pad=12)
    plt.tight_layout()

    path = os.path.join(save_dir, "03_CV_heatmap.png")
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def plot_dose_response_overlays(all_results: dict, dates: list, rep_colors: dict,
                                 save_dir: str) -> str:
    """
    For each antibiotic × strain: overlay dose-response curves from all biological replicates.
    Each replicate coloured by date; thin lines = technical reps, thick line = mean 4PL fit.
    """
    curves_dir   = os.path.join(save_dir, "dose_response_curves")
    os.makedirs(curves_dir, exist_ok=True)
    sorted_dates = sorted(dates)

    for strain, date_results in all_results.items():
        all_abs = sorted({r["ab_name"] for d in sorted_dates for r in date_results.get(d, [])})
        for ab_name in all_abs:
            fig, ax = plt.subplots(figsize=(8, 6))
            legend_handles = []
            for date in sorted_dates:
                color  = rep_colors[date]
                ab_res = next(
                    (r for r in date_results.get(date, []) if r["ab_name"] == ab_name), None
                )
                if ab_res is None:
                    continue
                conc = ab_res["concentrations"]
                xfit = np.logspace(np.log10(conc.min()), np.log10(conc.max()), 300)

                fit_curves = []
                for rep in ab_res["rep_results"]:
                    ax.scatter(conc, rep["responses"], color=color, alpha=0.4, s=25, zorder=4)
                    if rep["popt"] is not None:
                        yf = four_param_logistic(xfit, *rep["popt"])
                        ax.plot(xfit, yf, color=color, lw=1.2, alpha=0.4)
                        fit_curves.append(yf)

                if fit_curves:
                    ymean = np.nanmean(fit_curves, axis=0)
                    ystd  = np.nanstd(fit_curves, axis=0)
                    line, = ax.plot(xfit, ymean, color=color, lw=2.5, label=date)
                    ax.fill_between(xfit, ymean - ystd, ymean + ystd,
                                    color=color, alpha=0.15)
                    legend_handles.append(line)

                if not np.isnan(ab_res["mean_ic50"]):
                    ax.axvline(ab_res["mean_ic50"], color=color, ls=":", lw=1.5)

            ax.set_xscale("log")
            ax.set_xlabel("Concentration (µg/mL)", fontsize=12)
            ax.set_ylabel("Relative growth (%)", fontsize=12)
            ax.set_ylim(bottom=0)
            ax.set_title(f"{ab_name} – {strain}", fontsize=13, fontweight="bold")
            ax.grid(True, which="both", ls="--", lw=0.6, alpha=0.6)
            ax.legend(handles=legend_handles, title="Biological replicate", fontsize=9)
            plt.tight_layout()

            fname = f"{strain}_{ab_name}".replace(" ", "_").replace("/", "-") + ".png"
            fig.savefig(os.path.join(curves_dir, fname), dpi=300, bbox_inches="tight")
            plt.close(fig)

    return curves_dir


def _shade_color(hex_color: str, factor: float) -> str:
    """Blend hex_color toward white. factor=1.0 → original, factor=0.0 → white."""
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    return "#{:02x}{:02x}{:02x}".format(
        int(r * factor + 255 * (1 - factor)),
        int(g * factor + 255 * (1 - factor)),
        int(b * factor + 255 * (1 - factor)),
    )


def plot_ic50_full_summary(all_results: dict, strain: str, file_num: int,
                           rep_colors: dict, save_dir: str) -> str:
    """
    One bar per antibiotic × biological replicate × technical replicate.
    Biological replicates are colour-coded by date; technical replicates within
    each date are shown as light→dark shades of that date colour.
    Legend shows biological replicate dates; shade gradient indicates tech reps.
    """
    if strain not in all_results:
        return ""

    date_results = all_results[strain]
    sorted_dates = sorted(date_results.keys())
    all_abs = sorted({r["ab_name"] for d in sorted_dates for r in date_results.get(d, [])})
    if not all_abs:
        return ""

    n_ab     = len(all_abs)
    n_dates  = len(sorted_dates)
    n_tech   = 3
    bar_w    = 0.8 / (n_dates * n_tech)
    x        = np.arange(n_ab)
    # shade factors: lightest → darkest for tech rep 1, 2, 3
    shades   = [0.35, 0.65, 1.0]

    fig, ax = plt.subplots(figsize=(max(14, n_ab * 1.4), 6))

    legend_handles = []
    for di, date in enumerate(sorted_dates):
        base = rep_colors[date]
        colors = [_shade_color(base, f) for f in shades]

        res_lookup = {r["ab_name"]: r for r in date_results.get(date, [])}

        for ti in range(n_tech):
            # offset centres: (date group centre) + (tech rep offset within group)
            group_offset = (di - n_dates / 2 + 0.5) * (n_tech * bar_w)
            tech_offset  = (ti - n_tech  / 2 + 0.5) * bar_w
            offset = group_offset + tech_offset

            for ai, ab in enumerate(all_abs):
                res = res_lookup.get(ab)
                if res is None:
                    continue
                rep_ic50s = res["rep_ic50s"]
                if ti >= len(rep_ic50s) or np.isnan(rep_ic50s[ti]):
                    continue
                ax.bar(x[ai] + offset, rep_ic50s[ti], bar_w * 0.92,
                       color=colors[ti], zorder=3)

        legend_handles.append(Patch(facecolor=base, label=date))

    ax.set_yscale("log")
    ax.set_xticks(x)
    ax.set_xticklabels(all_abs, rotation=45, ha="right", fontsize=9)
    ax.set_ylabel("IC50 (µg/mL)", fontsize=12)
    ax.set_title(f"IC50 – {strain}  (shades = technical replicates within each day)",
                 fontsize=12, fontweight="bold")
    ax.legend(handles=legend_handles, title="Biological replicate",
              fontsize=9, title_fontsize=9)
    ax.grid(True, which="both", axis="y", ls="--", lw=0.5, alpha=0.5)
    ax.set_axisbelow(True)
    plt.tight_layout()

    fname = f"{file_num:02d}_IC50_full_summary_{strain.replace('-', '')}.png"
    path = os.path.join(save_dir, fname)
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


# ─────────────────────────────────────────────────────────────
# Variance component decomposition
# ─────────────────────────────────────────────────────────────
def compute_variance_components(all_results: dict) -> dict:
    """
    One-way ANOVA on log10(IC50) per (strain, antibiotic).
    Groups = biological replicates (dates); observations = technical row-replicates.

    Variance components (on the log10 scale):
      sigma2_tech = MS_within              (row-to-row noise within one day)
      sigma2_bio  = max(0, (MS_between - MS_within) / n_harm)  (day-to-day noise)

    CV% = (10^sigma - 1) * 100 converts log-scale SD back to interpretable units.
    ICC = sigma2_bio / (sigma2_bio + sigma2_tech) expresses how much of total variance
    is attributable to genuine biological differences between days (0=none, 1=all).

    Returns {strain: DataFrame} with one row per antibiotic.
    """
    strain_vc = {}
    for strain, date_results in all_results.items():
        dates   = sorted(date_results.keys())
        all_abs = sorted({r["ab_name"] for d in dates for r in date_results.get(d, [])})
        rows = []
        for ab_name in all_abs:
            groups = []
            for date in dates:
                ab_res = next(
                    (r for r in date_results.get(date, []) if r["ab_name"] == ab_name), None
                )
                if ab_res is None:
                    continue
                vals = [np.log10(v) for v in ab_res["rep_ic50s"] if not np.isnan(v) and v > 0]
                if vals:
                    groups.append(np.array(vals))

            row = {"Antibiotic": ab_name}
            if len(groups) < 2:
                rows.append(row)
                continue

            grand_mean = np.concatenate(groups).mean()
            k          = len(groups)
            n_harm     = k / sum(1.0 / len(g) for g in groups)

            SS_between = sum(len(g) * (g.mean() - grand_mean) ** 2 for g in groups)
            SS_within  = sum(np.sum((g - g.mean()) ** 2) for g in groups)
            df_between = k - 1
            df_within  = sum(len(g) - 1 for g in groups)

            MS_between = SS_between / df_between if df_between > 0 else np.nan
            MS_within  = SS_within  / df_within  if df_within  > 0 else np.nan

            if np.isnan(MS_between) or np.isnan(MS_within):
                rows.append(row)
                continue

            sigma2_tech = MS_within
            sigma2_bio  = max(0.0, (MS_between - MS_within) / n_harm)
            sigma2_total = sigma2_tech + sigma2_bio

            F_stat  = MS_between / MS_within if MS_within > 0 else np.nan
            p_anova = float(1 - stats.f.cdf(F_stat, df_between, df_within)) if not np.isnan(F_stat) else np.nan

            ICC       = sigma2_bio  / sigma2_total if sigma2_total > 0 else np.nan
            tech_frac = sigma2_tech / sigma2_total * 100 if sigma2_total > 0 else np.nan
            bio_frac  = sigma2_bio  / sigma2_total * 100 if sigma2_total > 0 else np.nan

            sigma_tech = np.sqrt(sigma2_tech)
            sigma_bio  = np.sqrt(sigma2_bio)
            cv_tech    = (10 ** sigma_tech - 1) * 100
            cv_bio     = (10 ** sigma_bio  - 1) * 100

            row.update({
                "CV_tech_pct":   cv_tech,
                "CV_bio_pct":    cv_bio,
                "tech_frac_pct": tech_frac,
                "bio_frac_pct":  bio_frac,
                "sigma_tech_log": sigma_tech,
                "sigma_bio_log":  sigma_bio,
                "ICC":           ICC,
                "F_stat":        F_stat,
                "p_anova":       p_anova,
            })
            rows.append(row)
        strain_vc[strain] = pd.DataFrame(rows)
    return strain_vc


def plot_variance_fractions(vc_dfs: dict, save_dir: str) -> str:
    """
    Stacked horizontal bar showing what fraction of total log-scale variance is
    technical vs biological for each antibiotic × strain combination.
    Immediately reveals which antibiotics are limited by technical vs biological noise.
    """
    fig_rows = []
    for strain, df in vc_dfs.items():
        valid = df.dropna(subset=["tech_frac_pct", "bio_frac_pct"]).copy()
        for _, row in valid.iterrows():
            fig_rows.append({
                "label":     f"{row['Antibiotic']} ({strain})",
                "tech_frac": row["tech_frac_pct"],
                "bio_frac":  row["bio_frac_pct"],
            })

    if not fig_rows:
        return ""

    labels    = [r["label"]     for r in fig_rows]
    tech_vals = [r["tech_frac"] for r in fig_rows]
    bio_vals  = [r["bio_frac"]  for r in fig_rows]
    y = np.arange(len(labels))

    fig, ax = plt.subplots(figsize=(10, max(6, len(labels) * 0.42)))
    ax.barh(y, tech_vals, color=_TECH_COLOR, alpha=0.85,
            label="Technical (within-day)", zorder=3)
    ax.barh(y, bio_vals, left=tech_vals, color=_BIO_COLOR, alpha=0.85,
            label="Biological (day-to-day)", zorder=3)
    ax.axvline(50, color="black", lw=1, ls="--", alpha=0.4)

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=8)
    ax.set_xlabel("% of total variance", fontsize=12)
    ax.set_xlim(0, 100)
    ax.set_title("Variance Partitioning: Technical vs Biological",
                 fontsize=13, fontweight="bold")
    ax.legend(fontsize=10, loc="lower right")
    ax.grid(True, axis="x", ls="--", lw=0.5, alpha=0.5)
    ax.set_axisbelow(True)
    plt.tight_layout()

    path = os.path.join(save_dir, "04_variance_fractions.png")
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def build_raw_ic50_df(all_results: dict, dates: list) -> pd.DataFrame:
    """
    Return a tidy DataFrame: one row per (antibiotic, strain, date).
    Columns: Antibiotic, Strain, Date, Rep1_IC50, Rep2_IC50, Rep3_IC50, Mean_IC50, SD_IC50, Fold_change
    Fold_change = max/min across dates for that antibiotic × strain.
    """
    rows = []
    for strain, date_results in all_results.items():
        for date in sorted(dates):
            for r in date_results.get(date, []):
                rep_vals = (list(r["rep_ic50s"]) + [np.nan, np.nan, np.nan])[:3]
                rows.append({
                    "Antibiotic": r["ab_name"],
                    "Strain":     strain,
                    "Date":       date,
                    "Rep1_IC50":  rep_vals[0],
                    "Rep2_IC50":  rep_vals[1],
                    "Rep3_IC50":  rep_vals[2],
                    "Mean_IC50":  r["mean_ic50"],
                    "SD_IC50":    r["sd_ic50"],
                })
    df = pd.DataFrame(rows)
    def fold_change(grp):
        means = grp["Mean_IC50"].dropna()
        if len(means) < 2:
            return pd.Series([np.nan] * len(grp), index=grp.index)
        return pd.Series([means.max() / means.min()] * len(grp), index=grp.index)
    df["Fold_change"] = df.groupby(["Antibiotic", "Strain"], group_keys=False).apply(fold_change)
    return df.sort_values(["Antibiotic", "Strain", "Date"]).reset_index(drop=True)


def flag_outlier_days(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each (Antibiotic, Strain), compute each date's fold from the central IC50
    (exp of mean log-IC50) across all dates. Flag: OK <1.5×, WARNING 1.5–2×, OUTLIER >2×.
    """
    rows = []
    for (ab, strain), grp in raw_df.groupby(["Antibiotic", "Strain"]):
        vals = grp["Mean_IC50"].dropna().values
        if len(vals) < 2 or np.any(vals <= 0):
            central_ic50 = np.nan
        else:
            central_ic50 = np.exp(np.mean(np.log(vals)))
        for _, row in grp.iterrows():
            ic50 = row["Mean_IC50"]
            fold = ic50 / central_ic50 if (not np.isnan(central_ic50) and ic50 > 0) else np.nan
            if np.isnan(fold):
                flag = "N/A"
            elif fold > 2.0 or fold < 0.5:
                flag = "OUTLIER"
            elif fold > 1.5 or fold < 0.667:
                flag = "WARNING"
            else:
                flag = "OK"
            rows.append({
                "Antibiotic":   ab,
                "Strain":       strain,
                "Date":         row["Date"],
                "Mean_IC50":    ic50,
                "FoldFromMean": fold,
                "Flag":         flag,
            })
    return pd.DataFrame(rows).sort_values(["Antibiotic", "Strain", "Date"]).reset_index(drop=True)


def plot_outlier_heatmap(outlier_df: pd.DataFrame, save_dir: str) -> str:
    """
    Colour-coded heatmap of fold-from-mean per antibiotic × date, faceted by strain.
    Green = OK, orange = WARNING, red = OUTLIER.
    """
    strains      = sorted(outlier_df["Strain"].unique())
    antibiotics  = sorted(outlier_df["Antibiotic"].unique())
    dates        = sorted(outlier_df["Date"].unique())
    n_s, n_ab, n_d = len(strains), len(antibiotics), len(dates)

    flag_colors = {"OK": "#a8d5a2", "WARNING": "#f9c74f", "OUTLIER": "#e63946", "N/A": "#dddddd"}

    fig, axes = plt.subplots(1, n_s, figsize=(6 * n_s, max(5, n_ab * 0.55 + 1.5)),
                              squeeze=False)
    fig.suptitle("Outlier Day Detection — Fold from Mean IC50 per Antibiotic",
                 fontsize=13, fontweight="bold")

    for si, strain in enumerate(strains):
        ax  = axes[0][si]
        sub = outlier_df[outlier_df["Strain"] == strain].copy()
        ax.set_xlim(-0.5, n_d - 0.5)
        ax.set_ylim(-0.5, n_ab - 0.5)

        for di, date in enumerate(dates):
            for ai, ab in enumerate(antibiotics):
                mask = (sub["Antibiotic"] == ab) & (sub["Date"] == date)
                row  = sub[mask]
                if row.empty:
                    flag, fold = "N/A", np.nan
                else:
                    flag = row.iloc[0]["Flag"]
                    fold = row.iloc[0]["FoldFromMean"]
                color = flag_colors.get(flag, "#dddddd")
                rect  = plt.Rectangle((di - 0.5, ai - 0.5), 1, 1,
                                       facecolor=color, edgecolor="white", lw=1.0)
                ax.add_patch(rect)
                label = f"{fold:.2g}×" if not np.isnan(fold) else "—"
                ax.text(di, ai, label, ha="center", va="center", fontsize=8,
                        color="black" if flag != "OUTLIER" else "white", fontweight="bold")

        ax.set_xticks(range(n_d))
        ax.set_xticklabels(dates, rotation=30, ha="right", fontsize=9)
        ax.set_yticks(range(n_ab))
        ax.set_yticklabels(antibiotics, fontsize=9)
        ax.set_title(f"Strain: {strain}", fontsize=11, fontweight="bold")
        ax.set_xlabel("Biological replicate date", fontsize=9)

    flag_labels = {
        "OK":      "OK  (fold < 1.5×)",
        "WARNING": "WARNING  (1.5× – 2×)",
        "OUTLIER": "OUTLIER  (> 2×)",
        "N/A":     "N/A",
    }
    legend_elements = [Patch(facecolor=flag_colors[f], label=flag_labels[f])
                       for f in ("OK", "WARNING", "OUTLIER", "N/A")]
    fig.legend(handles=legend_elements, loc="lower center", ncol=4,
               fontsize=9, framealpha=0.9, bbox_to_anchor=(0.5, -0.02))

    plt.tight_layout()
    path = os.path.join(save_dir, "06_outlier_flag_heatmap.png")
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    return path




def save_raw_ic50_excel(raw_df: pd.DataFrame, save_dir: str) -> str:
    """Export raw per-replicate IC50 table as a standalone Excel file."""
    path = os.path.join(save_dir, "05_raw_ic50_table.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        raw_df.rename(columns=_RAW_IC50_RENAME).to_excel(
            writer, index=False, sheet_name="Raw IC50 Values"
        )
        _autofit_columns(writer.sheets["Raw IC50 Values"], max_width=28)
    return path


# ─────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────
def _autofit_columns(ws, max_width: int = 32) -> None:
    for col in ws.columns:
        w = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(w, max_width)


def save_reproducibility_excel(strain_dfs: dict, dates: list, save_dir: str,
                               vc_dfs: dict = None, raw_df: pd.DataFrame = None,
                               outlier_df: pd.DataFrame = None,
                               ratio_df: pd.DataFrame = None) -> str:
    sorted_dates = sorted(dates)
    path         = os.path.join(save_dir, "IC50_reproducibility.xlsx")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Summary sheet: one row per strain
        summary_rows = []
        for strain, df in strain_dfs.items():
            valid = df.dropna(subset=["CV_pct"])
            summary_rows.append({
                "Strain":                    strain,
                "Antibiotics measured":      len(df),
                "Median CV%":                round(valid["CV_pct"].median(), 1) if not valid.empty else np.nan,
                "Mean CV%":                  round(valid["CV_pct"].mean(),   1) if not valid.empty else np.nan,
                "Max CV%":                   round(valid["CV_pct"].max(),    1) if not valid.empty else np.nan,
                "Antibiotics with CV < 20%": int((valid["CV_pct"] < 20).sum()),
            })
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Summary")

        # Per-strain detailed sheets
        for strain, df in strain_dfs.items():
            cols = ["Antibiotic"]
            for d in sorted_dates:
                for suffix in ["_IC50", "_SD"]:
                    c = f"{d}{suffix}"
                    if c in df.columns:
                        cols.append(c)
            cols += ["Mean_IC50", "SD_IC50", "CV_pct", "log10_SD", "n_reps"]
            cols = [c for c in cols if c in df.columns]
            col_rename = {
                "Mean_IC50": "Mean IC50 (µg/mL)",
                "SD_IC50":   "SD IC50 (µg/mL)",
                "CV_pct":    "CV%",
                "log10_SD":  "SD of log10(IC50)",
                "n_reps":    "N replicates",
                **{f"{d}_IC50": f"{d} IC50 (µg/mL)" for d in sorted_dates},
                **{f"{d}_SD":   f"{d} SD"            for d in sorted_dates},
            }
            sheet_name = strain[:31]
            df[cols].rename(columns=col_rename).to_excel(
                writer, index=False, sheet_name=sheet_name
            )
            _autofit_columns(writer.sheets[sheet_name], max_width=42)

        # Variance components sheet
        if vc_dfs:
            vc_rows = []
            for strain, df in vc_dfs.items():
                valid = df.dropna(subset=["CV_tech_pct"])
                for _, row in valid.iterrows():
                    vc_rows.append({
                        "Strain":                   strain,
                        "Antibiotic":               row["Antibiotic"],
                        "Technical CV%":            round(row["CV_tech_pct"],  1),
                        "Biological CV%":           round(row["CV_bio_pct"],   1),
                        "% Variance Technical":     round(row["tech_frac_pct"], 1),
                        "% Variance Biological":    round(row["bio_frac_pct"],  1),
                        "ICC":                      round(row["ICC"],           3),
                        "F (ANOVA)":                round(row["F_stat"],        3),
                        "p (ANOVA)":                row["p_anova"],
                        "sigma_tech (log10)":       round(row["sigma_tech_log"], 4),
                        "sigma_bio (log10)":        round(row["sigma_bio_log"],  4),
                    })
            vc_export = pd.DataFrame(vc_rows)
            vc_export.to_excel(writer, index=False, sheet_name="Variance Components")
            _autofit_columns(writer.sheets["Variance Components"], max_width=32)

        # Raw IC50 values sheet
        if raw_df is not None and not raw_df.empty:
            raw_df.rename(columns=_RAW_IC50_RENAME).to_excel(
                writer, index=False, sheet_name="Raw IC50 Values"
            )
            _autofit_columns(writer.sheets["Raw IC50 Values"], max_width=28)

        # Outlier flags sheet (colour-coded)
        if outlier_df is not None and not outlier_df.empty:
            outlier_export = outlier_df.rename(columns={
                "Mean_IC50":   "Mean IC50 (µg/mL)",
                "FoldFromMean": "Fold from mean",
            })
            outlier_export.to_excel(writer, index=False, sheet_name="Outlier Flags")
            ws  = writer.sheets["Outlier Flags"]
            fills = {
                "OK":      PatternFill("solid", fgColor="A8D5A2"),
                "WARNING": PatternFill("solid", fgColor="F9C74F"),
                "OUTLIER": PatternFill("solid", fgColor="E63946"),
                "N/A":     PatternFill("solid", fgColor="DDDDDD"),
            }
            flag_col_idx = outlier_export.columns.get_loc("Flag") + 1  # 1-based
            for row_idx, flag in enumerate(outlier_export["Flag"], start=2):
                cell = ws.cell(row=row_idx, column=flag_col_idx)
                cell.fill = fills.get(flag, fills["N/A"])
            _autofit_columns(ws, max_width=24)

        # Strain ratio sheet
        if ratio_df is not None and not ratio_df.empty:
            ratio_df.rename(columns={
                "ACE1_IC50":   "ACE-1 mean IC50 (µg/mL)",
                "MG1655_IC50": "MG1655 mean IC50 (µg/mL)",
                "Ratio":       "ACE-1 / MG1655 ratio",
            }).to_excel(writer, index=False, sheet_name="Strain Ratio")
            _autofit_columns(writer.sheets["Strain Ratio"], max_width=26)

    return path


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    plate_map = load_plate_map(PLATE_MAP_PATH)

    print("=" * 65)
    print("  IC50 Biological Replicate Comparison Pipeline")
    print("=" * 65)
    print(f"  Data dir : {BASE_DIR}")
    print(f"  Plate map: {PLATE_MAP_PATH}")
    print(f"  Output   : {OUTPUT_DIR}")
    print()

    file_groups = discover_files(BASE_DIR)
    strains     = sorted(file_groups.keys())
    all_dates   = sorted({d for s in file_groups.values() for d in s})
    rep_colors  = make_rep_colors(all_dates)

    print(f"  Strains found       : {', '.join(strains)}")
    print(f"  Biological replicates: {', '.join(all_dates)}")
    print()

    # Process all (strain, date) experiments
    all_results: dict = {}
    for strain in strains:
        all_results[strain] = {}
        for date in sorted(file_groups[strain]):
            print(f"[{strain}  |  {date}]")
            res = process_experiment(file_groups[strain][date], plate_map)
            all_results[strain][date] = res
            print(f"  -> {len(res)} antibiotics processed")
            for r in res:
                ic50_str = f"{r['mean_ic50']:.4g}" if not np.isnan(r["mean_ic50"]) else "N/A"
                print(f"     {r['ab_name']:<35}  IC50 = {ic50_str} µg/mL"
                      f"  techSD = {r['sd_ic50']:.4g}")
        print()

    # Build reproducibility tables
    print("[Reproducibility summary]")
    strain_dfs = build_reproducibility_dfs(all_results)
    for strain, df in strain_dfs.items():
        valid = df.dropna(subset=["CV_pct"])
        med   = valid["CV_pct"].median() if not valid.empty else np.nan
        mean  = valid["CV_pct"].mean()   if not valid.empty else np.nan
        n_ok  = int((valid["CV_pct"] < 20).sum()) if not valid.empty else 0
        print(f"  {strain:<10}  median CV = {med:.1f}%  mean CV = {mean:.1f}%  "
              f"({n_ok}/{len(valid)} antibiotics with CV < 20%)")
    print()

    # Variance component analysis
    print("[Variance component analysis]")
    vc_dfs = compute_variance_components(all_results)
    for strain, df in vc_dfs.items():
        valid = df.dropna(subset=["CV_tech_pct", "CV_bio_pct"])
        if not valid.empty:
            print(f"  {strain:<10}  tech CV = {valid['CV_tech_pct'].median():.1f}% (median)"
                  f"  bio CV = {valid['CV_bio_pct'].median():.1f}% (median)"
                  f"  ICC = {valid['ICC'].median():.3f} (median)")
    print()

    # Plots
    print("[1/7] Dose-response curve overlays (all bio replicates)...")
    p = plot_dose_response_overlays(all_results, all_dates, rep_colors, OUTPUT_DIR)
    print(f"      Saved in: {p}")

    for step, strain in [(2, "MG1655"), (3, "ACE-1")]:
        print(f"[{step}/7] IC50 full summary – {strain} (tech + bio replicates)...")
        p = plot_ic50_full_summary(all_results, strain, step, rep_colors, OUTPUT_DIR)
        if p:
            print(f"      Saved: {os.path.basename(p)}")

    print("[4/7] CV% heatmap...")
    p = plot_cv_heatmap(strain_dfs, OUTPUT_DIR)
    print(f"      Saved: {os.path.basename(p)}")

    print("[5/7] Variance fraction stacked bar...")
    p = plot_variance_fractions(vc_dfs, OUTPUT_DIR)
    print(f"      Saved: {os.path.basename(p)}")

    print("[6/7] Raw IC50 table (Excel)...")
    raw_df = build_raw_ic50_df(all_results, all_dates)
    p = save_raw_ic50_excel(raw_df, OUTPUT_DIR)
    print(f"      Saved: {os.path.basename(p)}")

    print("[7/7] Outlier day flagging heatmap...")
    outlier_df = flag_outlier_days(raw_df)
    p = plot_outlier_heatmap(outlier_df, OUTPUT_DIR)
    print(f"      Saved: {os.path.basename(p)}")

    print()
    print("Exporting Excel...")
    p = save_reproducibility_excel(strain_dfs, all_dates, OUTPUT_DIR,
                                   vc_dfs=vc_dfs, raw_df=raw_df,
                                   outlier_df=outlier_df)
    print(f"  Saved: {os.path.basename(p)}")

    print()
    print("=" * 65)
    print("  Done.")
    print(f"  Output folder: {OUTPUT_DIR}")
    print("=" * 65)


if __name__ == "__main__":
    main()
